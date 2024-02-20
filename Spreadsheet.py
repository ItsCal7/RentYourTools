from openpyxl import load_workbook

workbook = load_workbook("Database.xlsx")
userSheet = workbook["Users"]
toolSheet = workbook["Tools"]
lendSheet = workbook["Lends"]

def addAccount(account):
    userSheet.append([account.getUsername(), account.getPassword(), account.getAccountType()])
    workbook.save("Database.xlsx")

def deleteAccount(rowNum):
    userSheet.delete_rows(rowNum, 1)
    workbook.save("Database.xlsx")

def deleteLend(rowNum):
    lendSheet.delete_rows(rowNum, 1)
    workbook.save("Database.xlsx")

def addTool(tool):
    toolSheet.append([tool.getID(), tool.getPrice(), tool.getSize(), tool.getCondition(), tool.getAvailability()])
    workbook.save("Database.xlsx")

def addLend(lend, row):
    lendSheet.append([lend.getRenter(), lend.getReturnDate(), lend.getRentedTool(), lend.getPrice(), lend.checkPending()])
    toolSheet.cell(row, 5).value = False
    workbook.save("Database.xlsx")

def getValue(row, col, sheet):
    if sheet == "users":
        return userSheet.cell(row, col).value
    elif sheet == "tools":
        return toolSheet.cell(row, col).value
    elif sheet == "lends":
        return lendSheet.cell(row, col).value

def getRows(sheet):
    if sheet == "users":
        return userSheet.max_row
    elif sheet == "tools":
        return toolSheet.max_row
    elif sheet == "lends":
        return lendSheet.max_row

def returnTool(row):
    lendSheet.cell(row, 5).value = True
    workbook.save("Database.xlsx")

def confirmReturn(row):
    toolSheet.cell(row, 5).value = True
    workbook.save("Database.xlsx")

